# ============================================================
# IMPORTS
# ============================================================

import sys
import os
import json
import platform
import subprocess
import pandas as pd
import openpyxl
import zipfile
import tempfile
import re
from datetime import datetime
try:
    from docx import Document as _Document
except ImportError:
    _Document = None  # Labels feature unavailable; caught at runtime
from PyQt6.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QPushButton, QTextEdit,
    QFileDialog, QHBoxLayout, QVBoxLayout, QGroupBox, QMessageBox,
    QDialog, QTableWidget, QTableWidgetItem, QDialogButtonBox, QProgressBar,
    QComboBox
)
from PyQt6.QtGui import QFont, QIcon, QColor  # QColor needed for table cell backgrounds
from PyQt6.QtCore import Qt, QSize, QThread, pyqtSignal


# ============================================================
# ICON LOADING
# ============================================================

APP_DIR = getattr(__import__("sys"), "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
ICON_DIR = os.path.join(APP_DIR, "icons")

# Settings file stored beside the app so it persists across working-directory changes.
SETTINGS_FILE = os.path.join(APP_DIR, "gui_settings.json")


def load_icon(name: str) -> QIcon:
    """Load an icon from the /icons folder."""
    path = os.path.join(ICON_DIR, name)
    if os.path.exists(path):
        return QIcon(path)
    return QIcon()


# ============================================================
# XML NORMALIZATION FIX (MERGE SAME-FORMATTING RUNS ONLY)
# ============================================================

def normalize_xml_runs(xml_content: str) -> str:
    """
    Normalize Word XML so placeholders are not split across runs.
    Merges adjacent <w:t> tags within a run, and merges adjacent <w:r>
    tags only when they share identical run properties (<w:rPr>), so that
    bold, italic, font changes, etc. are preserved.

    """
    # Step 1: merge adjacent <w:t> tags within the same run
    xml_content = re.sub(r'</w:t>\s*<w:t[^>]*>', '', xml_content)

    # Step 2: merge adjacent <w:r> tags only when run properties are identical
    # Pattern: </w:r><w:r> where both runs have the same <w:rPr>...</w:rPr> block
    # (or both have no <w:rPr> at all).
    #
    # We do this by collapsing the closing tag + opening tag + duplicate rPr block:
    #   </w:r>  <w:r>  <w:rPr>SAME</w:rPr>  <w:t  →  <w:t
    # where SAME means the rPr content already present in the preceding run.
    xml_content = re.sub(
        r'(<w:rPr>)(.*?)(</w:rPr>)(.*?)</w:r>\s*<w:r>\s*\1\2\3\s*<w:t',
        r'\1\2\3\4<w:t',
        xml_content,
        flags=re.DOTALL
    )

    # Step 3: merge adjacent runs that both have NO run properties at all
    xml_content = re.sub(
        r'</w:r>\s*<w:r>\s*<w:t',
        r'<w:t',
        xml_content
    )

    return xml_content


# ============================================================
# HELPER — FORMAT CELL VALUES CLEANLY
# ============================================================

def format_cell_value(value) -> str:
    """
    Convert Excel cell values to clean strings.
    - Dates/datetimes → DD/MM/YYYY.
    - Whole-number floats → integer string (no decimal point).
    - NaN/None → empty string.
    - Lists/arrays are str()-converted directly to avoid pd.isna() ValueError.
    """
    # pd.isna() raises ValueError on collections — check type first.
    if isinstance(value, (list, dict)) or (
        hasattr(value, '__len__') and not isinstance(value, (str, bytes, pd.Timestamp))
    ):
        return str(value)
    try:
        if pd.isna(value):
            return ""
    except (ValueError, TypeError):
        # Unexpected type — treat as non-null and convert to string.
        return str(value)
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


# ============================================================
# HELPER — SPLIT ADDRESS
# ============================================================

def split_address(address_str: str) -> dict:
    """Split a comma-separated address string into Address1–Address4."""
    parts = [p.strip() for p in str(address_str).split(",")]
    while len(parts) < 4:
        parts.append("")
    if len(parts) > 4:
        parts = parts[:3] + [", ".join(parts[3:])]
    return {
        "Address1": parts[0],
        "Address2": parts[1],
        "Address3": parts[2],
        "Address4": parts[3],
    }


# ============================================================
# BACKEND — WORD TEMPLATE FILLER
# ============================================================

def replace_placeholders_in_xml(xml_content, replacements):
    for key, value in replacements.items():
        xml_content = xml_content.replace(key, value)
    return xml_content


def _build_combined_docx(entries, combined_path):
    """
    Combine multiple .docx files into one by working at the zip level.
    This preserves headers, footers, images, and all relationships that
    python-docx's element cloning approach silently drops.

    Strategy:
    - Use the first letter as the base document (retains its styles/settings/rels)
    - For each subsequent letter, import its body XML with a page break separator,
      and copy its media files into the combined archive under renamed paths to
      avoid collisions, updating rId references accordingly.
    """
    import shutil, uuid
    from lxml import etree

    WNS  = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    RONS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

    # Work in a temp dir — copy first letter as the starting base
    with tempfile.TemporaryDirectory() as work_dir:
        base_docx = os.path.join(work_dir, "combined_work.docx")
        shutil.copy2(entries[0]["file"], base_docx)

        # Extract base
        base_dir = os.path.join(work_dir, "base")
        with zipfile.ZipFile(base_docx, "r") as z:
            z.extractall(base_dir)

        # Parse base document body
        doc_xml_path = os.path.join(base_dir, "word", "document.xml")
        tree = etree.parse(doc_xml_path)
        root = tree.getroot()
        ns = {"w": WNS}
        body = root.find(".//w:body", ns)

        # Remove the trailing sectPr so we can re-append it after all content
        sect_pr = body.find("w:sectPr", ns)
        if sect_pr is not None:
            body.remove(sect_pr)

        # Track highest existing rId in base to avoid collisions when merging
        rels_path = os.path.join(base_dir, "word", "_rels", "document.xml.rels")
        rels_tree = etree.parse(rels_path)
        rels_root = rels_tree.getroot()
        existing_ids = {
            rel.get("Id") for rel in rels_root.findall("{%s}Relationship" % RONS)
        }

        def _next_rid(existing):
            i = 1
            while f"rId{i}" in existing:
                i += 1
            existing.add(f"rId{i}")
            return f"rId{i}"

        def _make_section_break(src_sect_pr):
            """
            Build a paragraph whose pPr contains the source sectPr.
            Word treats this as a continuous section boundary that starts
            a new page via the sectPr's own page definition, avoiding the
            double-page that an explicit <w:br type='page'/> causes when
            the template already fills a full page.
            """
            from copy import deepcopy
            p = etree.fromstring(
                b'<w:p xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
                b'<w:pPr/></w:p>'
            )
            pPr = p.find("{%s}pPr" % WNS)
            if src_sect_pr is not None:
                pPr.append(deepcopy(src_sect_pr))
            return p

        # Append each subsequent letter
        for entry in entries[1:]:
            with tempfile.TemporaryDirectory() as src_dir:
                with zipfile.ZipFile(entry["file"], "r") as z:
                    z.extractall(src_dir)

                # Build a rId remap for this source document
                src_rels_path = os.path.join(src_dir, "word", "_rels", "document.xml.rels")
                remap = {}  # old rId -> new rId
                if os.path.exists(src_rels_path):
                    src_rels_tree = etree.parse(src_rels_path)
                    src_rels_root = src_rels_tree.getroot()
                    for rel in src_rels_root.findall("{%s}Relationship" % RONS):
                        old_id  = rel.get("Id")
                        old_target = rel.get("Target", "")
                        rel_type = rel.get("Type", "")

                        # Copy media files with a unique name to avoid collisions
                        if old_target.startswith("media/"):
                            src_media = os.path.join(src_dir, "word", old_target)
                            if os.path.exists(src_media):
                                ext = os.path.splitext(old_target)[1]
                                new_name = f"media/{uuid.uuid4().hex}{ext}"
                                dst_media = os.path.join(base_dir, "word", new_name)
                                os.makedirs(os.path.dirname(dst_media), exist_ok=True)
                                shutil.copy2(src_media, dst_media)
                                new_id = _next_rid(existing_ids)
                                remap[old_id] = new_id
                                new_rel = etree.SubElement(
                                    rels_root,
                                    "{%s}Relationship" % RONS,
                                    Id=new_id, Type=rel_type, Target=new_name
                                )

                # Parse source body and remap rId attributes
                src_doc_path = os.path.join(src_dir, "word", "document.xml")
                src_tree = etree.parse(src_doc_path)
                src_root = src_tree.getroot()
                src_body = src_root.find(".//w:body", ns)

                if src_body is not None:
                    src_sect_pr = src_body.find("{%s}sectPr" % WNS)
                    elements = [el for el in list(src_body) if not el.tag.endswith("}sectPr")]
                    for el in elements:
                        # Remap any r:id / r:embed attributes
                        for node in el.iter():
                            for attr in list(node.attrib):
                                if attr.endswith("}id") or attr.endswith("}embed") or attr.endswith("}href"):
                                    val = node.get(attr)
                                    if val in remap:
                                        node.set(attr, remap[val])
                        body.append(el)
                    # Add section break paragraph — carries the source sectPr so
                    # Word starts a new page without inserting a blank extra page
                    body.append(_make_section_break(src_sect_pr))

        # Re-append sectPr at the very end
        if sect_pr is not None:
            body.append(sect_pr)

        # Write updated rels
        rels_tree.write(rels_path, xml_declaration=True, encoding="UTF-8", standalone=True)

        # Write updated document.xml
        tree.write(doc_xml_path, xml_declaration=True, encoding="UTF-8", standalone=True)

        # Repack as a new zip
        with zipfile.ZipFile(combined_path, "w", zipfile.ZIP_DEFLATED) as out_zip:
            for dirpath, _, filenames in os.walk(base_dir):
                for fname in filenames:
                    fpath = os.path.join(dirpath, fname)
                    arcname = os.path.relpath(fpath, base_dir)
                    out_zip.write(fpath, arcname)



def _extract_template_placeholders(template_path):
    """
    Extract all {{ColumnName}} placeholders from a .docx template.
    Returns a set of column name strings (without the braces).
    Address placeholders (Address1-Address4) are collapsed to "Address".
    """
    ADDRESS_PLACEHOLDERS = {"Address1", "Address2", "Address3", "Address4"}
    found = set()
    xml_files = [
        "word/document.xml",
        "word/header1.xml", "word/header2.xml", "word/header3.xml",
        "word/footer1.xml", "word/footer2.xml", "word/footer3.xml",
    ]
    with zipfile.ZipFile(template_path, "r") as z:
        names = z.namelist()
        for xml_file in xml_files:
            if xml_file in names:
                text = z.read(xml_file).decode("utf-8", errors="ignore")
                for match in re.finditer(r"\{\{([^}]+)\}\}", text):
                    key = match.group(1).strip()
                    if key in ADDRESS_PLACEHOLDERS:
                        found.add("Address")
                    else:
                        found.add(key)
    return found


def fill_word_template(excel_path, template_path, output_dir, progress_cb=None):
    # Validate the template is a real .docx (zip) before extracting.
    if not zipfile.is_zipfile(template_path):
        raise ValueError(
            f"The Word template does not appear to be a valid .docx file:\n{template_path}"
        )

    df = pd.read_excel(excel_path)
    df.columns = [str(c).strip() for c in df.columns]

    if "Letter_Generated" not in df.columns:
        df["Letter_Generated"] = ""

    # Cast to object dtype so string values can be assigned later.
    # An all-blank column is inferred as float64 by pandas, which rejects strings.
    df["Letter_Generated"] = df["Letter_Generated"].astype(object).where(
        df["Letter_Generated"].notna(), ""
    )

    # Create 'Print Docs' subfolder inside the chosen output folder
    print_docs_dir = os.path.join(output_dir, "Print Docs")
    os.makedirs(print_docs_dir, exist_ok=True)

    created_files = []

    pending_rows = [(i, r) for i, r in df.iterrows()
                    if str(r["Letter_Generated"]).strip().lower() != "yes"]
    total = len(pending_rows)

    for step, (index, row) in enumerate(pending_rows, start=1):
        if progress_cb:
            row_name = str(row["Name"]).strip() if "Name" in df.columns else f"Row {index + 1}"
            progress_cb(step, total, row_name)

        address_map = {}
        if "Address" in df.columns:
            address_map = split_address(row["Address"])

        # Build the placeholder→value map for this row.
        replacements = {}
        for col in df.columns:
            replacements[f"{{{{{col}}}}}"] = format_cell_value(row[col])

        for key, value in address_map.items():
            replacements[f"{{{{{key}}}}}"] = value

        # Apply any column mappings the user defined via the mapping dialog.
        mapping_path = template_path + ".colmap.json"
        if os.path.exists(mapping_path):
            with open(mapping_path, "r") as _f:
                _colmap = json.load(_f)
            for placeholder, col_name in _colmap.items():
                if col_name and col_name in df.columns:
                    replacements[f"{{{{{placeholder}}}}}"] = format_cell_value(row[col_name])
                elif col_name is None:
                    replacements[f"{{{{{placeholder}}}}}"] = ""

        # Derive output filename from the Name column; fall back to row number.
        if "Name" in df.columns:
            base_name = str(row["Name"]).replace("/", "-")
        else:
            base_name = f"Document_{index + 1}"

        output_path = os.path.join(output_dir, f"{base_name}.docx")

        # Append a counter suffix if a file with this name already exists.
        if os.path.exists(output_path):
            counter = 1
            while True:
                candidate = os.path.join(output_dir, f"{base_name}_{counter}.docx")
                if not os.path.exists(candidate):
                    output_path = candidate
                    break
                counter += 1

        # Extract the template zip, perform replacements, repack as a new .docx.
        with tempfile.TemporaryDirectory() as tmpdir:
            with zipfile.ZipFile(template_path, 'r') as zip_ref:
                zip_ref.extractall(tmpdir)

            xml_files = [
                "word/document.xml",
                "word/header1.xml",
                "word/header2.xml",
                "word/header3.xml",
                "word/footer1.xml",
                "word/footer2.xml",
                "word/footer3.xml",
            ]

            for xml_file in xml_files:
                full_path = os.path.join(tmpdir, xml_file)
                if os.path.exists(full_path):

                    # Detect encoding from the XML declaration; fall back to utf-8.
                    with open(full_path, "rb") as f:
                        raw = f.read()

                    encoding = "utf-8"
                    match = re.search(rb'encoding=["\']([^"\']+)["\']', raw[:200])
                    if match:
                        encoding = match.group(1).decode("ascii")

                    xml_content = raw.decode(encoding)

                    # Normalise split runs so placeholders aren't fragmented across XML runs.
                    xml_content = normalize_xml_runs(xml_content)

                    xml_content = replace_placeholders_in_xml(xml_content, replacements)

                    with open(full_path, "w", encoding=encoding) as f:
                        f.write(xml_content)

            # Repack the modified XML back into a .docx archive.
            with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as new_zip:
                for folder, _, files in os.walk(tmpdir):
                    for file in files:
                        file_path = os.path.join(folder, file)
                        arcname = os.path.relpath(file_path, tmpdir)
                        new_zip.write(file_path, arcname)

        df.at[index, "Letter_Generated"] = "yes"

        created_files.append({
            "name": str(row["Name"]) if "Name" in df.columns else f"Row {index + 1}",
            "file": output_path
        })

    # Write Letter_Generated status back to the Excel file, patching only the
    # changed cells so existing formatting and formulas in the workbook are preserved.
    if created_files:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # Locate the Letter_Generated column in the sheet header row
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
        col_index = None
        for cell in header_row:
            if str(cell.value).strip() == "Letter_Generated":
                col_index = cell.column
                break

        if col_index is None:
            # Column doesn't exist yet — append it to the header
            col_index = ws.max_column + 1
            ws.cell(row=1, column=col_index, value="Letter_Generated")

        # Write back only the rows that were updated (df index == Excel row - 1,
        # offset by 2 to account for the 1-based row index and the header row).
        for df_index, row_value in df["Letter_Generated"].items():
            excel_row = df_index + 2  # +1 for 1-based, +1 for header
            ws.cell(row=excel_row, column=col_index, value=row_value)

        wb.save(excel_path)

    # Build combined.docx at the zip level so headers, images, and all
    # relationships are preserved exactly as they appear in each individual letter.
    if created_files:
        combined_path = os.path.join(print_docs_dir, "combined.docx")
        _build_combined_docx(created_files, combined_path)
        created_files.append({"name": "Combined Document", "file": combined_path})

    return created_files


# ============================================================
# LABEL PRODUCTION  — Avery L7163  (2 col × 7 row, 14 labels per A4 sheet)
# Label: 99.1 × 38.1 mm  |  H-gap: 2 mm  |  V-gap: 0 mm
# Top/bottom margin: 15.15 mm  |  Left/right: 4.9 mm
# ============================================================

def generate_labels_doc(excel_path, output_folder, start_slot=1):
    if _Document is None:
        raise RuntimeError("python-docx is not installed. Run: pip install python-docx")

    from lxml import etree

    # ── Exact L7163 dimensions in twips (1 twip = 1/1440 inch = 25.4/1440 mm) ──
    # Page margins are set so the text area is EXACTLY 7 × row_h tall.
    # This is how Word's own label wizard avoids the 7th-row pagination problem.
    def t(mm):
        return str(int(round(mm * 1440 / 25.4)))

    PAGE_W  = t(210);   PAGE_H  = t(297)
    MAR_TOP = t(15.15); MAR_BOT = t(15.15)
    MAR_L   = t(4.9);   MAR_R   = t(4.9)
    COL_W   = t(99.1)   # label column width
    GAP_W   = t(2.0)    # spacer column between the two label columns
    ROW_H   = t(38.1)   # label row height
    COLS    = 2
    ROWS    = 7

    W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"

    def el(tag, **kw):
        n = etree.Element(f"{{{W}}}{tag}")
        for k, v in kw.items():
            n.set(f"{{{W}}}{k}", str(v))
        return n

    def ch(p, tag, **kw):
        n = el(tag, **kw)
        p.append(n)
        return n

    # ── Read data ──────────────────────────────────────────────────────────────
    df = pd.read_excel(excel_path)
    df.columns = [str(c).strip() for c in df.columns]
    if df.empty:
        raise RuntimeError("Excel file has no rows.")

    labels = []
    for _, row in df.iterrows():
        name = str(row["Name"]).strip() if "Name" in df.columns else ""
        amap = {}
        if "Address" in df.columns:
            raw = row["Address"]
            if raw and not (isinstance(raw, float) and pd.isna(raw)):
                amap = split_address(str(raw))
        lines = [l for l in [name,
            amap.get("Address1",""), amap.get("Address2",""),
            amap.get("Address3",""), amap.get("Address4",""),
        ] if l.strip()]
        labels.append(lines)

    if not any(labels):
        raise RuntimeError("No label content — check Name/Address columns.")

    # Insert empty slots at the front so the first label lands in the correct cell.
    offset = max(0, start_slot - 1)
    labels = [[] for _ in range(offset)] + labels

    per_page = COLS * ROWS
    rem = len(labels) % per_page
    if rem:
        labels += [[] for _ in range(per_page - rem)]

    # ── Build document ─────────────────────────────────────────────────────────
    doc = _Document()

    # Page size and margins — set via sectPr so Word knows the exact text area
    sec = doc.sections[0]
    from docx.oxml import OxmlElement
    from docx.oxml.ns import qn

    # Set page size
    sectPr = sec._sectPr
    pgSz = sectPr.find(qn("w:pgSz"))
    if pgSz is None:
        pgSz = OxmlElement("w:pgSz")
        sectPr.append(pgSz)
    pgSz.set(qn("w:w"), PAGE_W)
    pgSz.set(qn("w:h"), PAGE_H)

    # Set margins — bottom = 15.15mm so text area = exactly 7 × 38.1mm
    pgMar = sectPr.find(qn("w:pgMar"))
    if pgMar is None:
        pgMar = OxmlElement("w:pgMar")
        sectPr.append(pgMar)
    pgMar.set(qn("w:top"),    MAR_TOP)
    pgMar.set(qn("w:bottom"), MAR_BOT)
    pgMar.set(qn("w:left"),   MAR_L)
    pgMar.set(qn("w:right"),  MAR_R)
    pgMar.set(qn("w:header"), "0")
    pgMar.set(qn("w:footer"), "0")
    pgMar.set(qn("w:gutter"), "0")

    # Remove default empty paragraph
    body = doc.element.body
    for p in list(body):
        if p.tag != f"{{{W}}}sectPr":
            body.remove(p)

    total_pages = len(labels) // per_page

    for page_num in range(total_pages):

        # ── Table ──────────────────────────────────────────────────────────────
        # No explicit page break needed — each table exactly fills the text area
        # (7 rows × 2160 twips = 15120 = page height − top margin − bottom margin)
        # so Word naturally starts each subsequent table on a new page.
        tbl = el("tbl")
        body.insert(list(body).index(body.find(f"{{{W}}}sectPr")), tbl)

        # tblPr
        tblPr = ch(tbl, "tblPr")
        ch(tblPr, "tblW",          w=str(int(COL_W)*COLS + int(GAP_W)*(COLS-1)), type="dxa")
        ch(tblPr, "tblInd",        w=MAR_L, type="dxa")   # position table at left margin
        ch(tblPr, "tblCellSpacing",w="0",  type="dxa")
        ch(tblPr, "tblLayout",     type="fixed")
        tblBd = ch(tblPr, "tblBorders")
        for side in ("top","left","bottom","right","insideH","insideV"):
            ch(tblBd, side, val="none", sz="0", space="0", color="auto")
        tblMar = ch(tblPr, "tblCellMar")
        for side in ("top","start","bottom","end"):
            ch(tblMar, side, w="0", type="dxa")

        # tblGrid: label col, gap col, label col
        tblGrid = ch(tbl, "tblGrid")
        ch(tblGrid, "gridCol", w=COL_W)
        ch(tblGrid, "gridCol", w=GAP_W)
        ch(tblGrid, "gridCol", w=COL_W)

        for r in range(ROWS):
            tr = ch(tbl, "tr")
            trPr = ch(tr, "trPr")
            ch(trPr, "trHeight", val=ROW_H, hRule="exact")
            ch(trPr, "cantSplit")

            for c in range(COLS):
                # Label cell
                tc = ch(tr, "tc")
                tcPr = ch(tc, "tcPr")
                ch(tcPr, "tcW", w=COL_W, type="dxa")
                tcBd = ch(tcPr, "tcBorders")
                for side in ("top","left","bottom","right"):
                    ch(tcBd, side, val="none", sz="0", space="0", color="auto")
                tcMar = ch(tcPr, "tcMar")
                ch(tcMar, "top",   w="113", type="dxa")   # ~2mm top padding
                ch(tcMar, "start", w="1134", type="dxa")  # 2cm left indent
                ch(tcMar, "bottom",w="0",   type="dxa")
                ch(tcMar, "end",   w="0",   type="dxa")

                slot = page_num * per_page + r * COLS + c
                lines = labels[slot]

                if lines:
                    for line in lines:
                        p = ch(tc, "p")
                        pPr = ch(p, "pPr")
                        ch(pPr, "spacing", before="0", after="0",
                           line="240", lineRule="exact")
                        ch(pPr, "jc", val="left")
                        rn = ch(p, "r")
                        rPr = ch(rn, "rPr")
                        ch(rPr, "sz",   val="20")  # 10pt
                        ch(rPr, "szCs", val="20")
                        t_el = ch(rn, "t")
                        t_el.text = line
                        if line != line.strip():
                            t_el.set(
                                "{http://www.w3.org/XML/1998/namespace}space",
                                "preserve"
                            )
                else:
                    ch(tc, "p")

                # Gap spacer cell (not after last column)
                if c < COLS - 1:
                    gc = ch(tr, "tc")
                    gcPr = ch(gc, "tcPr")
                    ch(gcPr, "tcW", w=GAP_W, type="dxa")
                    gcBd = ch(gcPr, "tcBorders")
                    for side in ("top","left","bottom","right"):
                        ch(gcBd, side, val="none", sz="0", space="0", color="auto")
                    ch(gc, "p")

    # ── Save ───────────────────────────────────────────────────────────────────
    print_docs_dir = os.path.join(output_folder, "Print Docs")
    os.makedirs(print_docs_dir, exist_ok=True)
    output_path = os.path.join(print_docs_dir, "labels.docx")
    doc.save(output_path)

    if not os.path.exists(output_path):
        raise RuntimeError(f"labels.docx failed to save at:\n{output_path}")

    return output_path

# ============================================================
# PREVIEW WINDOW
# ============================================================

class PreviewWindow(QDialog):
    def __init__(self, excel_path):
        super().__init__()
        self.setWindowTitle("Preview Rows")
        self.resize(900, 500)

        layout = QVBoxLayout()
        self.setLayout(layout)

        # Guard data loading so file errors show a message rather than crashing.
        try:
            df = pd.read_excel(excel_path)
        except Exception as e:
            layout.addWidget(QLabel(f"Could not load Excel file:\n{e}"))
            button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
            button_box.rejected.connect(self.reject)
            layout.addWidget(button_box)
            return

        df.columns = [str(c).strip() for c in df.columns]

        # Add the column if the file has never been processed.
        if "Letter_Generated" not in df.columns:
            df["Letter_Generated"] = ""

        # Reset index so row positions are always contiguous.
        df = df.reset_index(drop=True)


        label = QLabel(
            "Green rows will have letters generated.  Red rows have already been processed."
        )
        layout.addWidget(label)

        table = QTableWidget()
        layout.addWidget(table)

        table.setRowCount(len(df))
        table.setColumnCount(len(df.columns))
        table.setHorizontalHeaderLabels(list(df.columns))

        # Locate Letter_Generated by name rather than position.
        lg_col_pos = df.columns.get_loc("Letter_Generated")

        for r, (_, row) in enumerate(df.iterrows()):
            for c, col in enumerate(df.columns):
                raw = row[col]
                # Show empty string for NaN cells rather than "nan".
                if pd.isna(raw) if not isinstance(raw, (list, dict)) else False:
                    display = ""
                else:
                    display = str(raw)
                table.setItem(r, c, QTableWidgetItem(display))

            already_done = str(row.iloc[lg_col_pos]).strip().lower() == "yes"

            # Colour rows: green = will generate, red = already done.
            # setForeground is required — PyQt6 does not auto-contrast text colour.
            bg = QColor(255, 180, 180) if already_done else QColor(180, 255, 180)
            fg = QColor(0, 0, 0)  # black text on both backgrounds
            for c in range(len(df.columns)):
                table.item(r, c).setBackground(bg)
                table.item(r, c).setForeground(fg)

        table.resizeColumnsToContents()

        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)


# ============================================================
# SETTINGS
# ============================================================

_SETTINGS_DEFAULTS = {"excel": "", "word": "", "output": "", "label_next_slot": "1"}

def load_settings():
    # Load settings, merging saved values over defaults.
    # Merging rather than validating means old settings files missing newer keys
    # (e.g. label_next_slot) still load correctly instead of falling back to empty.
    result = dict(_SETTINGS_DEFAULTS)
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r") as f:
                data = json.load(f)
            if isinstance(data, dict):
                for k in _SETTINGS_DEFAULTS:
                    if k in data and isinstance(data[k], str):
                        result[k] = data[k]
        except (json.JSONDecodeError, OSError):
            pass
    return result

def save_settings(excel, word, output, label_next_slot="1"):
    with open(SETTINGS_FILE, "w") as f:
        json.dump({"excel": excel, "word": word, "output": output, "label_next_slot": str(label_next_slot)}, f)



# ============================================================
# LABEL START DIALOG
# ============================================================

class LabelStartDialog(QDialog):
    """
    Shown before generating labels. Displays the remembered next label
    slot and lets the user override it. Slots are numbered 1–14 left to
    right across the sheet then down (Avery L7163 layout).
    """
    def __init__(self, remembered_slot: int, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Label Start Position")
        self.resize(360, 160)

        layout = QVBoxLayout()
        self.setLayout(layout)

        info = QLabel(
            f"Labels will start from slot {remembered_slot} "
            f"(1\u201314, left to right across the sheet then down).\n"
            f"Change the value below if you are using a partially used sheet."
        )
        info.setWordWrap(True)
        layout.addWidget(info)

        row = QWidget()
        row_layout = QHBoxLayout()
        row_layout.setContentsMargins(0, 8, 0, 0)
        row.setLayout(row_layout)

        row_layout.addWidget(QLabel("Start from slot:"))
        self._spin = QComboBox()
        self._spin.addItems([str(i) for i in range(1, 15)])
        self._spin.setCurrentIndex(remembered_slot - 1)
        row_layout.addWidget(self._spin)
        row_layout.addStretch()
        layout.addWidget(row)

        btn_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btn_box.accepted.connect(self.accept)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def start_slot(self) -> int:
        return int(self._spin.currentText())


# ============================================================
# COLUMN MAPPING DIALOG
# ============================================================

class ColumnMappingDialog(QDialog):
    """
    Shown when the template contains placeholders that don't match
    any Excel column. The user picks which column maps to each one,
    or skips it (leaves it blank in the output).
    """
    def __init__(self, missing_placeholders, excel_columns, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Column Mapping Required")
        self.resize(500, 400)
        self._combos = {}

        layout = QVBoxLayout()
        self.setLayout(layout)

        info = QLabel(
            "The following template placeholders could not be matched to "
            "an Excel column.\nSelect the correct column for each, or choose "
            "\'(skip — leave blank)\' to ignore it."
        )
        info.setWordWrap(True)
        layout.addWidget(info)

        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout()
        scroll_widget.setLayout(scroll_layout)

        options = ["(skip — leave blank)"] + sorted(excel_columns)

        for placeholder in sorted(missing_placeholders):
            row_widget = QWidget()
            row_layout = QHBoxLayout()
            row_layout.setContentsMargins(0, 0, 0, 0)
            row_widget.setLayout(row_layout)

            lbl = QLabel(f"{{{{{placeholder}}}}}")
            lbl.setMinimumWidth(180)
            lbl.setStyleSheet("font-family: monospace; font-weight: bold;")
            combo = QComboBox()
            combo.addItems(options)

            # Auto-select if a case-insensitive match exists
            for i, opt in enumerate(options):
                if opt.lower() == placeholder.lower():
                    combo.setCurrentIndex(i)
                    break

            row_layout.addWidget(lbl)
            row_layout.addWidget(combo, 1)
            scroll_layout.addWidget(row_widget)
            self._combos[placeholder] = combo

        scroll_layout.addStretch()
        layout.addWidget(scroll_widget)

        btn_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btn_box.accepted.connect(self.accept)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def get_mapping(self):
        """
        Returns dict of {placeholder: excel_column_or_None}.
        None means the user chose to skip it.
        """
        result = {}
        for placeholder, combo in self._combos.items():
            chosen = combo.currentText()
            result[placeholder] = None if chosen.startswith("(skip") else chosen
        return result


# ============================================================
# BACKGROUND WORKERS
# ============================================================

class RunWorker(QThread):
    """Runs fill_word_template in a background thread. Returns a list of dicts."""
    finished = pyqtSignal(list)
    progress = pyqtSignal(int, int, str)  # (current, total, name)
    error    = pyqtSignal(str)

    def __init__(self, fn, *args):
        super().__init__()
        self._fn, self._args = fn, args

    def run(self):
        try:
            result = self._fn(*self._args, progress_cb=self.progress.emit)
            self.finished.emit(result)
        except Exception as e:
            self.error.emit(str(e))


class LabelsWorker(QThread):
    """Runs generate_labels_doc in a background thread. Returns output path or empty string."""
    finished = pyqtSignal(str)
    error    = pyqtSignal(str)

    def __init__(self, fn, *args):
        super().__init__()
        self._fn, self._args = fn, args

    def run(self):
        try:
            result = self._fn(*self._args)
            # pyqtSignal(str) cannot carry None — convert to empty string
            self.finished.emit(result if result is not None else "")
        except Exception as e:
            self.error.emit(str(e))


# ============================================================
# MAIN WINDOW
# ============================================================

class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowIcon(load_icon("app.png"))
        self.setWindowTitle("Karl's Letter Maker")
        self.resize(800, 600)

        self.settings = load_settings()
        # Separate slots for each worker — prevents GC collecting a running thread.
        self._run_worker    = None
        self._labels_worker = None
        self.init_ui()

    def button_icon(self, button: QPushButton, icon_filename: str):
        icon = load_icon(icon_filename)
        button.setIcon(icon)
        button.setIconSize(QSize(32, 32))

    def init_ui(self):
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(12)

        # Header row — large app icon on the left, title text on the right
        header_layout = QHBoxLayout()
        header_layout.setSpacing(14)

        app_icon_label = QLabel()
        app_icon = load_icon("app.png")
        if not app_icon.isNull():
            app_icon_label.setPixmap(app_icon.pixmap(QSize(196, 196)))
        app_icon_label.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        header_layout.addWidget(app_icon_label)

        title_label = QLabel("Karl's Letter Maker")
        title_label.setFont(QFont("Century Gothic", 32, QFont.Weight.Bold))
        title_label.setAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
        header_layout.addWidget(title_label, 1)

        main_layout.addLayout(header_layout)

        # Excel group
        excel_group = QGroupBox("Excel File")
        excel_layout = QHBoxLayout()
        self.excel_edit = QLineEdit(self.settings["excel"])
        browse_excel_btn = QPushButton("Browse")
        self.button_icon(browse_excel_btn, "excel.png")
        browse_excel_btn.clicked.connect(self.select_excel)
        excel_layout.addWidget(self.excel_edit)
        excel_layout.addWidget(browse_excel_btn)
        excel_group.setLayout(excel_layout)
        main_layout.addWidget(excel_group)

        # Word group
        word_group = QGroupBox("Word Template")
        word_layout = QHBoxLayout()
        self.word_edit = QLineEdit(self.settings["word"])
        browse_word_btn = QPushButton("Browse")
        self.button_icon(browse_word_btn, "word.png")
        browse_word_btn.clicked.connect(self.select_word)
        word_layout.addWidget(self.word_edit)
        word_layout.addWidget(browse_word_btn)
        word_group.setLayout(word_layout)
        main_layout.addWidget(word_group)

        # Output group
        output_group = QGroupBox("Output Folder")
        output_layout = QHBoxLayout()
        self.output_edit = QLineEdit(self.settings["output"])
        browse_output_btn = QPushButton("Browse")
        self.button_icon(browse_output_btn, "folder.png")
        browse_output_btn.clicked.connect(self.select_output_folder)
        output_layout.addWidget(self.output_edit)
        output_layout.addWidget(browse_output_btn)
        output_group.setLayout(output_layout)
        main_layout.addWidget(output_group)

        # Action buttons
        action_layout = QHBoxLayout()

        self.preview_btn = QPushButton("Preview Rows")
        self.button_icon(self.preview_btn, "preview.png")
        self.preview_btn.clicked.connect(self.preview_rows)

        self.run_btn = QPushButton("Run Process")
        self.button_icon(self.run_btn, "run.png")
        self.run_btn.clicked.connect(self.run_process)

        self.open_output_btn = QPushButton("Open Output Folder")
        self.button_icon(self.open_output_btn, "folder_open.png")
        self.open_output_btn.clicked.connect(self.open_output_folder)

        self.labels_btn = QPushButton("Generate Labels")
        self.button_icon(self.labels_btn, "clipboard.png")
        self.labels_btn.clicked.connect(self.generate_labels)

        action_layout.addWidget(self.preview_btn)
        action_layout.addWidget(self.run_btn)
        action_layout.addWidget(self.labels_btn)
        action_layout.addWidget(self.open_output_btn)

        main_layout.addLayout(action_layout)

        # Progress bar — hidden until an operation is running
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 0)        # indeterminate / pulse mode
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setFixedHeight(6)
        self.progress_bar.hide()
        main_layout.addWidget(self.progress_bar)

        # Status label — sits next to the progress bar
        self.status_label = QLabel("")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_label.hide()
        main_layout.addWidget(self.status_label)

        # Log window
        log_group = QGroupBox("Created Files Log")
        log_layout = QVBoxLayout()
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        log_layout.addWidget(self.log_text)
        log_group.setLayout(log_layout)
        main_layout.addWidget(log_group)

        self.setLayout(main_layout)

    # ---------- File pickers ----------
    def select_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xlsx *.xls)")
        if path:
            self.excel_edit.setText(path)
            self._save_current_settings()

    def select_word(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Word Template", "", "Word Documents (*.docx)")
        if path:
            self.word_edit.setText(path)
            self._save_current_settings()

    def select_output_folder(self):
        path = QFileDialog.getExistingDirectory(self, "Select Output Folder", "")
        if path:
            self.output_edit.setText(path)
            self._save_current_settings()

    def _save_current_settings(self):
        save_settings(
            self.excel_edit.text().strip(),
            self.word_edit.text().strip(),
            self.output_edit.text().strip(),
            self.settings.get("label_next_slot", "1")
        )

    # ---------- Preview ----------
    def preview_rows(self):
        excel_path = self.excel_edit.text().strip()

        if not excel_path or not os.path.exists(excel_path):
            QMessageBox.critical(self, "Error", "Excel file not found.")
            return

        # CRASH FIX: catch any unexpected errors from PreviewWindow construction
        try:
            preview = PreviewWindow(excel_path)
            preview.exec()
        except Exception as e:
            QMessageBox.critical(self, "Preview Error", f"Could not open preview:\n{e}")

    # ---------- Open output folder ----------
    def open_output_folder(self):
        folder = self.output_edit.text().strip()
        if not folder or not os.path.isdir(folder):
            QMessageBox.critical(self, "Error", "Output folder is invalid.")
            return

        system = platform.system()
        try:
            if system == "Windows":
                os.startfile(folder)
            elif system == "Darwin":
                subprocess.Popen(["open", folder])
            else:
                subprocess.Popen(["xdg-open", folder])
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Could not open folder:\n{e}")

    # ---------- Run backend ----------
    def run_process(self):
        excel_path = self.excel_edit.text().strip()
        word_path = self.word_edit.text().strip()
        output_path = self.output_edit.text().strip()

        if not excel_path or not word_path or not output_path:
            QMessageBox.critical(self, "Missing Information", "Please fill in all fields.")
            return

        if not os.path.exists(excel_path):
            QMessageBox.critical(self, "Error", "Excel file not found.")
            return

        if not os.path.exists(word_path):
            QMessageBox.critical(self, "Error", "Word template not found.")
            return

        if not os.path.isdir(output_path):
            QMessageBox.critical(self, "Error", "Output folder does not exist.")
            return

        save_settings(excel_path, word_path, output_path)

        # Compare template placeholders to Excel columns.
        # Show a mapping dialog for any that cannot be matched automatically.
        try:
            template_placeholders = _extract_template_placeholders(word_path)
            df_cols = pd.read_excel(excel_path, nrows=0).columns.tolist()
            df_cols = [str(c).strip() for c in df_cols]
            mapping_path = word_path + ".colmap.json"

            # Load any previously saved mapping from a prior run.
            saved_mapping = {}
            if os.path.exists(mapping_path):
                try:
                    with open(mapping_path, "r") as _f:
                        saved_mapping = json.load(_f)
                except Exception:
                    saved_mapping = {}

            covered = set(df_cols)
            if any(c.lower() == "address" for c in df_cols):
                covered.add("Address")
            covered.update(saved_mapping.keys())

            missing = template_placeholders - covered
            if missing:
                dlg = ColumnMappingDialog(missing, df_cols, parent=self)
                if dlg.exec() != QDialog.DialogCode.Accepted:
                    return
                saved_mapping.update(dlg.get_mapping())
                with open(mapping_path, "w") as _f:
                    json.dump(saved_mapping, _f)
        except Exception as e:
            QMessageBox.warning(self, "Mapping Check Failed",
                f"Could not check placeholder mapping:\n{e}\n\nProceeding anyway.")

        self.log_text.clear()
        self.run_btn.setEnabled(False)
        self.run_btn.setText("Running…")
        self.progress_bar.setRange(0, 0)   # indeterminate until row count known
        self.progress_bar.show()
        self.status_label.setText("Loading spreadsheet…")
        self.status_label.show()

        self._run_worker = RunWorker(fill_word_template, excel_path, word_path, output_path)
        self._run_worker.finished.connect(self._on_run_finished)
        self._run_worker.progress.connect(self._on_run_progress)
        self._run_worker.error.connect(self._on_run_error)
        self._run_worker.start()

    def _on_run_progress(self, current, total, name):
        if self.progress_bar.maximum() == 0:
            self.progress_bar.setRange(0, total)
        self.progress_bar.setValue(current)
        self.status_label.setText(f"Generating letter {current} of {total}…")
        label = name if name.startswith("Row ") else f"\"{name}\""
        self.log_text.append(f"Letter generated for {label}")

    def _on_run_finished(self, created_files):
        self.run_btn.setEnabled(True)
        self.run_btn.setText("Run Process")
        self.progress_bar.hide()
        self.status_label.hide()
        # Log the combined doc entry if one was created
        combined = [e for e in created_files if e["name"] == "Combined Document"]
        if combined:
            self.log_text.append("Combined document created")
        QMessageBox.information(
            self, "Success",
            f"Process completed.\n\n{len(created_files)} Word document(s) created."
        )

    def _on_run_error(self, message):
        self.run_btn.setEnabled(True)
        self.run_btn.setText("Run Process")
        self.progress_bar.hide()
        self.status_label.hide()
        QMessageBox.critical(self, "Backend Error", message)

    def generate_labels(self):
        excel_path = self.excel_edit.text().strip()
        output_path = self.output_edit.text().strip()

        if not excel_path or not os.path.exists(excel_path):
            QMessageBox.critical(self, "Error", "Excel file not found.")
            return

        if not output_path or not os.path.isdir(output_path):
            QMessageBox.critical(self, "Error", "Output folder is invalid.")
            return

        # Show the start-slot dialog, pre-filled with the last remembered position.
        remembered = int(self.settings.get("label_next_slot", "1"))
        dlg = LabelStartDialog(remembered, parent=self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        self._label_start_slot = dlg.start_slot()

        # Run in a background thread to keep the UI responsive.
        self.labels_btn.setEnabled(False)
        self.labels_btn.setText("Generating…")
        self.progress_bar.setRange(0, 0)   # indeterminate for labels
        self.progress_bar.show()
        self.status_label.setText("Generating labels document…")
        self.status_label.show()

        self._labels_worker = LabelsWorker(generate_labels_doc, excel_path, output_path, self._label_start_slot)
        self._labels_worker.finished.connect(self._on_labels_finished)
        self._labels_worker.error.connect(self._on_labels_error)
        self._labels_worker.start()

    def _on_labels_finished(self, result):
        self.labels_btn.setEnabled(True)
        self.labels_btn.setText("Generate Labels")
        self.progress_bar.hide()
        self.status_label.hide()

        if not result:
            QMessageBox.information(self, "No Labels", "All letters have already been generated.")
            return

        # Work out how many labels were generated so we know which slot comes next.
        # Read the Excel to count rows, add the start offset, wrap within 14.
        try:
            excel_path = self.excel_edit.text().strip()
            n_labels = len(pd.read_excel(excel_path))
            start = getattr(self, "_label_start_slot", 1)
            next_slot = ((start - 1 + n_labels) % 14) + 1
            self.settings["label_next_slot"] = str(next_slot)
            save_settings(
                self.settings["excel"],
                self.settings["word"],
                self.settings["output"],
                next_slot
            )
        except Exception:
            pass  # Non-critical — don't block the success message

        self.log_text.append("Labels document generated")
        QMessageBox.information(self, "Labels Created", f"Labels saved to:\n{result}\n\nUse 'Open Output Folder' to access the file.")

    def _on_labels_error(self, message):
        self.labels_btn.setEnabled(True)
        self.labels_btn.setText("Generate Labels")
        self.progress_bar.hide()
        self.status_label.hide()
        QMessageBox.critical(self, "Labels Error", f"Failed to generate labels:\n{message}")


# ============================================================
# ENTRY POINT
# ============================================================

if __name__ == "__main__":
    app = QApplication(sys.argv)
    # Use the native Windows style where available; Fusion is the fallback.
    from PyQt6.QtWidgets import QStyleFactory
    available = QStyleFactory.keys()
    for style_name in ("windowsvista", "Windows", "Fusion"):
        if style_name.lower() in [s.lower() for s in available]:
            app.setStyle(style_name)
            break
    # Set Century Gothic as the application-wide font; Qt falls back to the
    # closest available sans-serif if it is not installed.
    app.setFont(QFont("Century Gothic", 10))
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
